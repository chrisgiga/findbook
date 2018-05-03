# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import scrapy
from scrapy.pipelines.images import ImagesPipeline
from scrapy.utils.project import get_project_settings
import datetime
from openpyxl import Workbook

class JpBookCrawlerPipeline(object):
    def __init__(self):
         self.settings  = get_project_settings()
         self.filePath  = self.settings.get('FILEPATH')
         self.wb = Workbook() 
         self.ws = self.wb.active
         self.ws.append(['書名', '作者', '出版社', '圖片', '業種', '頁數', 'isbn', 'ean', '售價', '售價(台幣)', '類型', '發售日','簡介', 'url'])
    def process_item(self, item, spider):
        line = [ item['title_main'], item['author'], item["publisher"], item['picture'], item['category'], item['page'], item['isbn'], item['ean'], item['price'], item['price_twd'], item['type'], item['sale_date'], item['content'], item['url'] ]
        self.ws.append(line)
        self.wb.save(self.filePath+'/jp/'+datetime.datetime.now().strftime("%Y%m%d")+'.xlsx')

        return item

class ImagesPipline(ImagesPipeline):
    def get_media_requests(self, item, info):       
        for image_url in item['image_urls']:
            yield scrapy.Request(image_url, meta={'image_name': item["image_name"]+'.jpg'})
    
    def file_path(self, request, response=None, info=None):
        return request.meta['image_name']

class EnBookCrawlerPipeline(object):
    def __init__(self):
         self.settings  = get_project_settings()
         self.filePath  = self.settings.get('FILEPATH')
         self.wb = Workbook() 
         self.ws = self.wb.active
         self.ws.append(['書名', '作者', '發售日', '出版社', 'isbn', '售價(原始)', '售價(台幣)', '類型', '圖片','簡介', 'url'])
    def process_item(self, item, spider):
#        self.csvwriter.writerow([item['title_main'], item['author'], item['sale_date'], item['publisher'], item['isbn13'], item['price_org'], item['price_twd'], item['type'], item['picture'], item['content'], item['url']])
        line = [ item['title_main'], item['author'], item['sale_date'], item['publisher'], item['isbn13'], item['price_org'], item['price_twd'], item['category'], item['picture'], item['content'], item['url'] ]
        self.ws.append(line)
        self.wb.save(self.filePath+'/en/'+datetime.datetime.now().strftime("%Y%m%d")+'.xlsx')
        return item
    
class CnBookCrawlerPipeline(object):
    def __init__(self):
         self.settings  = get_project_settings()
         self.filePath  = self.settings.get('FILEPATH')
         self.wb = Workbook() 
         self.ws = self.wb.active
         self.ws.append(['書名', '作者', '出版社', '發售日', 'isbn', '售價(人民幣)', '售價(台幣)', '類型', '圖片', '簡介', 'url'])
    def process_item(self, item, spider):
        line = [item['title_main'], item['author'], item['publisher'], item['sale_date'], item['isbn'], item['price_cny'], item['price_twd'], item['type'], item['picture'], item['content'], item['url'] ]
        self.ws.append(line)
        self.wb.save(self.filePath+'/cn/'+datetime.datetime.now().strftime("%Y%m%d")+'.xlsx')

        return item